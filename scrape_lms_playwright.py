import re
import time
from difflib import SequenceMatcher

import pandas as pd
from bs4 import BeautifulSoup
from playwright.sync_api import sync_playwright

from openpyxl import load_workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter


COURSE_ID = 48
BASE_URL = f"https://lms.kitakerja.id/grade/report/grader/index.php?id={COURSE_ID}"
STORAGE_FILE = "storage_state.json"
MASTER_FILE = "master_peserta_b.csv"
OUTPUT_FILE = "nilai_tugas_lms_kelas_b.xlsx"
PASSING_GRADE = 70


TARGET_COLUMNS = [
    "No",
    "First name / Last name",
    "Email address",
    "Pre-test",
    "Assignment: Tugas Praktik_Prinsip Manajemen Akses (Real)",
    "Assignment: Tugas Praktik_Prinsip Enkripsi Hashing Windows Ubuntu Kali (Real)",
    "Assignment: Tugas-Check Domain 1 (Real)",
    "Assignment: Tugas Checkpoint Domain 2 (Real)",
    "Assignment: Tugas Checkpoint Domain 3 (Real)",
    "Assignment: Tugas Chapter 4 (Real)",
    "Assignment: Checkpoint Domain 5 (Real)",
    "Post-test",
    "Capstone Project/PBL",
    "nilai akhir",
    "Status Kelulusan",
]


def clean_text(text):
    if not text:
        return ""
    return re.sub(r"\s+", " ", text.replace("\xa0", " ")).strip()


def parse_number(value):
    value = clean_text(value)
    if not value or value == "-":
        return ""

    value = value.replace(",", ".")
    match = re.search(r"-?\d+(?:\.\d+)?", value)

    if not match:
        return ""

    return float(match.group(0))


def normalize_header(header):
    header = clean_text(header)
    header = header.replace("QuizPre-Test", "Quiz Pre-Test")
    header = header.replace("QuizPost-Test", "Quiz Post-Test")
    header = header.replace("NaturalCourse total", "Natural Course total")
    header = re.sub(r"^Assignment(?=\S)", "Assignment ", header)
    header = re.sub(r"^Quiz(?=\S)", "Quiz ", header)
    header = re.sub(r"^Natural(?=\S)", "Natural ", header)
    return clean_text(header)


def clean_header_for_match(text):
    text = str(text or "").lower()
    text = text.replace("_", " ")
    text = text.replace("-", " ")
    text = text.replace(":", " ")
    text = text.replace("(", " ")
    text = text.replace(")", " ")
    text = text.replace("[", " ")
    text = text.replace("]", " ")
    return re.sub(r"\s+", " ", text).strip()


def header_match(source_col, target_col):
    source = clean_header_for_match(source_col)

    if "deletion in progress" in source:
        return False

    mapping_keywords = {
        "Pre-test": [
            "quizpre test",
            "quiz pre test",
            "pre test",
        ],
        "Assignment: Tugas Praktik_Prinsip Manajemen Akses (Real)": [
            "assignmenttugas praktik prinsip manajemen akses",
            "assignment tugas praktik prinsip manajemen akses",
            "prinsip manajemen akses",
            "manajemen akses",
        ],
        "Assignment: Tugas Praktik_Prinsip Enkripsi Hashing Windows Ubuntu Kali (Real)": [
            "assignmenttugas praktik prinsip enkripsi hashing windows ubuntu kali",
            "assignment tugas praktik prinsip enkripsi hashing windows ubuntu kali",
            "prinsip enkripsi hashing windows ubuntu kali",
            "enkripsi hashing",
            "windows ubuntu kali",
        ],
        "Assignment: Tugas-Check Domain 1 (Real)": [
            "assignmenttugas check domain 1",
            "assignment tugas check domain 1",
            "tugas check domain 1",
            "check domain 1",
            "domain 1",
        ],
        "Assignment: Tugas Checkpoint Domain 2 (Real)": [
            "assignmenttugas checkpoint domain 2",
            "assignment tugas checkpoint domain 2",
            "tugas checkpoint domain 2",
            "checkpoint domain 2",
            "domain 2",
        ],
        "Assignment: Tugas Checkpoint Domain 3 (Real)": [
            "assignmenttugas checkpoint domain 3",
            "assignment tugas checkpoint domain 3",
            "tugas checkpoint domain 3",
            "checkpoint domain 3",
            "domain 3",
        ],
        "Assignment: Tugas Chapter 4 (Real)": [
            "assignmenttugas chapter 4",
            "assignment tugas chapter 4",
            "tugas chapter 4",
            "chapter 4",
        ],
        "Assignment: Checkpoint Domain 5 (Real)": [
            "assignmentcheckpoint domain 5",
            "assignment checkpoint domain 5",
            "checkpoint domain 5",
            "domain 5",
        ],
        "Post-test": [
            "quizpost test",
            "quiz post test",
            "post test",
        ],
        "Capstone Project/PBL": [
            "capstone project",
            "capstone",
            "pbl",
        ],
        "nilai akhir": [
            "naturalcourse total",
            "natural course total",
            "course total",
            "nilai akhir",
            "final grade",
            "total",
        ],
    }

    for keyword in mapping_keywords.get(target_col, []):
        if keyword in source:
            return True

    return False


def find_grade_table(soup):
    tables = soup.find_all("table")

    if not tables:
        raise RuntimeError("Tidak menemukan table gradebook.")

    best_table = None
    best_score = 0

    for table in tables:
        text = clean_text(table.get_text(" "))
        score = 0

        if "Email address" in text:
            score += 5
        if "Assignment" in text or "Tugas" in text:
            score += 3
        if "Course total" in text or "NaturalCourse total" in text:
            score += 3
        if re.search(r"[\w\.-]+@[\w\.-]+\.\w+", text):
            score += 5

        if score > best_score:
            best_score = score
            best_table = table

    if not best_table:
        raise RuntimeError("Table ditemukan, tetapi bukan table nilai.")

    return best_table


def extract_headers(table):
    rows = table.find_all("tr")
    best_headers = []

    for row in rows:
        cells = row.find_all(["th", "td"])
        texts = [normalize_header(c.get_text(" ")) for c in cells]
        joined = " ".join(texts)

        if "Email address" in joined or "First name" in joined:
            best_headers = texts

    if not best_headers and rows:
        best_headers = [normalize_header(c.get_text(" ")) for c in rows[0].find_all(["th", "td"])]

    return best_headers


def extract_rows_from_html(html):
    soup = BeautifulSoup(html, "html.parser")
    text = soup.get_text(" ")

    if "Log in" in text and "Password" in text:
        raise RuntimeError("Session belum login atau sudah expired. Jalankan ulang login_lms.py.")

    table = find_grade_table(soup)
    headers = extract_headers(table)

    rows = []
    email_regex = re.compile(r"[\w\.-]+@[\w\.-]+\.\w+")

    for tr in table.find_all("tr"):
        cells = tr.find_all(["th", "td"])
        if not cells:
            continue

        values = [clean_text(c.get_text(" ")) for c in cells]
        row_text = " ".join(values)

        email_match = email_regex.search(row_text)
        if not email_match:
            continue

        email = email_match.group(0)

        email_index = None
        for i, value in enumerate(values):
            if email in value:
                email_index = i
                break

        if email_index is None:
            continue

        name = ""
        for i in range(0, email_index):
            candidate = values[i]
            if not candidate:
                continue
            if len(candidate) <= 3 and candidate.isupper():
                continue
            if candidate.lower() in ["user", "select", "picture"]:
                continue
            name = candidate
            break

        row_data = {
            "First name / Last name": name,
            "Email address": email,
        }

        grade_headers = headers[email_index + 1:]
        grade_values = values[email_index + 1:]

        for idx, value in enumerate(grade_values):
            if idx < len(grade_headers):
                col_name = normalize_header(grade_headers[idx])
            else:
                col_name = f"Unknown Grade {idx + 1}"

            row_data[col_name] = parse_number(value)

        rows.append(row_data)

    return rows, soup


def detect_pages(soup):
    pages = {0}

    for a in soup.find_all("a", href=True):
        href = a["href"]
        match = re.search(r"page=(\d+)", href)
        if match:
            pages.add(int(match.group(1)))

    return sorted(pages)


def normalize_to_sheet_format(raw_df):
    final_rows = []

    for idx, row in raw_df.iterrows():
        output = {col: "" for col in TARGET_COLUMNS}
        output["No"] = idx + 1
        output["First name / Last name"] = row.get("First name / Last name", "")
        output["Email address"] = row.get("Email address", "")

        for target_col in TARGET_COLUMNS:
            if target_col in ["No", "First name / Last name", "Email address", "Status Kelulusan"]:
                continue

            for source_col in raw_df.columns:
                if header_match(source_col, target_col):
                    output[target_col] = row.get(source_col, "")
                    break

        final_grade = output.get("nilai akhir", "")

        if isinstance(final_grade, (int, float)):
            output["Status Kelulusan"] = "Lulus" if final_grade >= PASSING_GRADE else "Tidak Lulus"

        final_rows.append(output)

    return pd.DataFrame(final_rows, columns=TARGET_COLUMNS)


def normalize_name(name):
    name = str(name or "").lower().strip()
    name = re.sub(r"[^a-z0-9\s]", " ", name)
    return re.sub(r"\s+", " ", name).strip()


def normalize_email(email):
    return str(email or "").lower().strip()


def name_similarity(a, b):
    return SequenceMatcher(None, normalize_name(a), normalize_name(b)).ratio()


def validate_with_master(final_df, master_csv):
    master_df = pd.read_csv(master_csv).fillna("")

    lms_df = final_df.copy().fillna("")
    lms_df["email_norm"] = lms_df["Email address"].apply(normalize_email)
    lms_df["name_norm"] = lms_df["First name / Last name"].apply(normalize_name)

    checked = []

    for _, master in master_df.iterrows():
        master_name = master["Nama"]
        master_email = master["Email"]
        master_email_norm = normalize_email(master_email)
        master_name_norm = normalize_name(master_name)

        result = {
            "Nama Master": master_name,
            "Kelas": master["Kelas"],
            "Email Master": master_email,
            "Status LMS": "",
            "Matched By": "",
            "Nama LMS": "",
            "Email LMS": "",
            "Catatan": "",
        }

        matched = None

        if master_email_norm:
            email_match = lms_df[lms_df["email_norm"] == master_email_norm]
            if not email_match.empty:
                matched = email_match.iloc[0]
                result["Status LMS"] = "Ada"
                result["Matched By"] = "Email"

        if matched is None:
            name_match = lms_df[lms_df["name_norm"] == master_name_norm]
            if not name_match.empty:
                matched = name_match.iloc[0]
                result["Status LMS"] = "Ada"
                result["Matched By"] = "Nama Exact"

        if matched is None:
            best_score = 0
            best_row = None

            for _, lms_row in lms_df.iterrows():
                score = name_similarity(master_name, lms_row["First name / Last name"])
                if score > best_score:
                    best_score = score
                    best_row = lms_row

            if best_score >= 0.86:
                matched = best_row
                result["Status LMS"] = "Ada"
                result["Matched By"] = f"Nama Mirip ({best_score:.2f})"

        if matched is not None:
            result["Nama LMS"] = matched["First name / Last name"]
            result["Email LMS"] = matched["Email address"]

            if not master_email_norm:
                result["Catatan"] = "Email master kosong, peserta ditemukan berdasarkan nama."
            elif normalize_email(matched["Email address"]) != master_email_norm:
                result["Catatan"] = "Nama ditemukan, tetapi email berbeda."
            else:
                result["Catatan"] = "OK"
        else:
            result["Status LMS"] = "Tidak Ada"
            result["Matched By"] = "-"
            result["Catatan"] = "Nama/email dari master tidak ditemukan di hasil scrape LMS."

        checked.append(result)

    validation_df = pd.DataFrame(checked)

    extra_rows = []
    master_emails = set(master_df["Email"].apply(normalize_email))
    master_names = set(master_df["Nama"].apply(normalize_name))

    for _, lms_row in lms_df.iterrows():
        lms_email = normalize_email(lms_row["Email address"])
        lms_name = normalize_name(lms_row["First name / Last name"])

        if lms_email not in master_emails and lms_name not in master_names:
            extra_rows.append({
                "Nama LMS": lms_row["First name / Last name"],
                "Email LMS": lms_row["Email address"],
                "Catatan": "Ada di LMS, tetapi tidak ada di master peserta.",
            })

    return validation_df, pd.DataFrame(extra_rows)


def format_excel(output_file):
    wb = load_workbook(output_file)

    blue_fill = PatternFill("solid", fgColor="4285F4")
    white_font = Font(color="FFFFFF", bold=True)
    border = Border(
        left=Side(style="thin", color="000000"),
        right=Side(style="thin", color="000000"),
        top=Side(style="thin", color="000000"),
        bottom=Side(style="thin", color="000000"),
    )

    for ws in wb.worksheets:
        max_col = ws.max_column
        max_row = ws.max_row

        if ws.title == "Nilai Tugas":
            ws.merge_cells(start_row=1, start_column=1, end_row=1, end_column=max_col)
            ws.cell(row=1, column=1).value = "Nilai Tugas"
            header_rows = [1, 2]
            ws.freeze_panes = "D3"
        else:
            header_rows = [1]
            ws.freeze_panes = "A2"

        for row_num in header_rows:
            for cell in ws[row_num]:
                cell.fill = blue_fill
                cell.font = white_font
                cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
                cell.border = border

        for row in ws.iter_rows(min_row=max(header_rows) + 1, max_row=max_row, min_col=1, max_col=max_col):
            for cell in row:
                cell.alignment = Alignment(vertical="center", wrap_text=True)
                cell.border = border

        for col_idx in range(1, max_col + 1):
            ws.column_dimensions[get_column_letter(col_idx)].width = 25

        if ws.title == "Nilai Tugas":
            ws.row_dimensions[1].height = 24
            ws.row_dimensions[2].height = 80

    wb.save(output_file)


def main():
    all_rows = []
    seen_emails = set()

    with sync_playwright() as p:
        browser = p.chromium.launch(
            headless=False,
            channel="chromium"
        )

        context = browser.new_context(storage_state=STORAGE_FILE)
        page = context.new_page()

        print("[*] Buka page 0...")
        page.goto(f"{BASE_URL}&page=0", wait_until="networkidle")
        time.sleep(2)

        html = page.content()
        first_rows, soup = extract_rows_from_html(html)
        pages = detect_pages(soup)

        print(f"[*] Page terdeteksi: {pages}")

        for page_num in pages:
            print(f"[*] Scraping page {page_num}...")

            if page_num == 0:
                rows = first_rows
            else:
                page.goto(f"{BASE_URL}&page={page_num}", wait_until="networkidle")
                time.sleep(2)
                rows, _ = extract_rows_from_html(page.content())

            for row in rows:
                email = row.get("Email address")
                if email and email not in seen_emails:
                    seen_emails.add(email)
                    all_rows.append(row)

            time.sleep(2)

        browser.close()

    if not all_rows:
        raise RuntimeError("Tidak ada data LMS yang berhasil diambil.")

    raw_df = pd.DataFrame(all_rows)
    final_df = normalize_to_sheet_format(raw_df)
    validation_df, extra_lms_df = validate_with_master(final_df, MASTER_FILE)

    with pd.ExcelWriter(OUTPUT_FILE, engine="openpyxl") as writer:
        final_df.to_excel(writer, sheet_name="Nilai Tugas", index=False, startrow=1)
        validation_df.to_excel(writer, sheet_name="Validasi Peserta", index=False)
        extra_lms_df.to_excel(writer, sheet_name="Extra LMS", index=False)

    format_excel(OUTPUT_FILE)

    print("[+] Selesai.")
    print(f"[+] Total peserta LMS: {len(final_df)}")
    print(f"[+] Output: {OUTPUT_FILE}")


if __name__ == "__main__":
    main()